# ref:   https://openpyxl.readthedocs.io/en/stable/usage.html
# ref:   https://openpyxl.readthedocs.io/en/stable/tutorial.html
# larryVid:   https://zoom.us/rec/share/yZRwBYiz2GJJGbORsX3DUIgoB920X6a80CRM8_dbyUowTM1x9eFkMsUiR7xMRxpP

 

import openpyxl
from openpyxl import load_workbook
      # from openpyxl import Workbook
      # import xlrd


wb = load_workbook(filename = 'Comp230_excel.xlsx')
      #  sheet_ranges = wb['customer_id','name','address','email']
      #print(sheet_ranges['D18'].value)
for sheet in wb:
    print(sheet.title)  #iterate through 
    for row in sheet.values:
        for value in row:
            print(value)


#workbook = xlrd.open_workbook('Comp230_excel.xlsx')
#{'customer_id':worksheet.cell(2,0).value,'name':worksheet.cell(2,1).value,'address':worksheet.cell(2,2).value,'email':worksheet.cell(3,3).value}
# worksheet = workbook.sheet_by_name('invoices')
# worksheet = workbook.sheet_by_name('itemssold')
# worksheet = workbook.sheet_by_name('productlist')



# customers=[]
# customers={"customers":[]}
# invoices={}
# items_sold={}
# products={}
# print(rowa)
# customers.append(rowa)
# customers.append(rowb)
# print("hello World")
#

# {
#    [
#         {
#             "id": 1
#             "name": John
#             "address": 233
#             "email:":john@hotmail.com
#         }
#         {
#             "id": 2
#             "name": anne
#             "address": 2511
#             "email:":anne@hotmail.com
#         }
#     ]
# }