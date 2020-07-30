# ref:   https://openpyxl.readthedocs.io/en/stable/usage.html
# ref:   https://openpyxl.readthedocs.io/en/stable/tutorial.html
# larryVid:   https://zoom.us/rec/share/yZRwBYiz2GJJGbORsX3DUIgoB920X6a80CRM8_dbyUowTM1x9eFkMsUiR7xMRxpP
import openpyxl
from openpyxl import load_workbook


wb = load_workbook(filename = 'Comp230_excel.xlsx')

client_sheet = wb["clients"]
customers={}
def Customer_call():
      for row in client_sheet.iter_rows(min_row=2,min_col=0,values_only=True):
            customer_id = row[0]
            cust = {
                        "name": row[1],
                        "address": row[2],
                        "email": row[3]
                  }
            customers[customer_id] = cust
      return customers





# import openpyxl
# from openpyxl import load_workbook

# CRED ='\033[91m'
# CEND ='\033[0m'
# CGREEN = '\33[32m'
# CBLINK ='\33[5m'
 
# wb = load_workbook(filename = 'Comp230_excel.xlsx')

# client_sheet = wb["clients"]
# customers={}

# for row in client_sheet.iter_rows(min_row=2,min_col=0,values_only=True):
#       customer_id = row[0]
#       cust = {
#                   "name": row[1],
#                   "address": row[2],
#                   "email": row[3]
#             }
#       customers[customer_id] = cust

# invoice_sheet = wb["invoices"] 
# invoices={}
# for row in invoice_sheet.iter_rows(min_row=2,min_col=0,values_only=True):
#       invoice_id = row[0]
#       inv = {
#                   "customer_id": row[1],
#                   "date": row[2]  
#             }
#       invoices[invoice_id] = inv

# itemsold_sheet = wb["itemsold"] 
# itemsold={}
# for row in itemsold_sheet.iter_rows(min_row=2,min_col=0,values_only=True):
#       items_sold_id = row[0]
#       item = {
#                   "invoice_id": row[1],
#                   "product_id": row[2],
#                   "quantity_sold": row[3]
#             }
#       itemsold[items_sold_id] = item

# productlist_sheet = wb["productlist"] 
# productlist={}
# for row in productlist_sheet.iter_rows(min_row=2,min_col=0,values_only=True):
#       product_id = row[0]
#       pro = {
#                   "product_name": row[1],
#                   "wieght": row[2],
#                   "price": row[3] 
#             }
#       productlist[product_id] = pro
# #print(productlist)
# # print(itemsold)
# print('Enter your invoice#')

# x = int(input())

# def invoice_printer():
#       date=invoices[x]['date']
#       invoiced_customer_id=invoices[x]['customer_id']
#       customer_name=customers[invoiced_customer_id]['name']
#       customer_email=customers[invoiced_customer_id]['email']
#       overall_price=0
#       print(CGREEN + f"Customer report for _{customer_name}_ regarding invoice #_{x}_ on {date} "+ CEND )
#       for y in itemsold: 
#             if (itemsold[y]['invoice_id'])==x:
#                   product_dict=itemsold[y]['product_id']
#                   product_name=productlist[product_dict]['product_name']
#                   price=productlist[product_dict]['price']
#                   # print(itemsold[y]['product_id'])
#                   sold=itemsold[y]['quantity_sold']
#                   total_price=round(sold*price,2)
#                   overall_price+=total_price
#                   print(f"Product:  {product_name} Quantity:  {sold} Price per unit:  ${price} Total Price:  ${total_price} ")
#       print(CRED + f"Total invoice pricing is: ${round(overall_price,2)}"+ CEND)
# invoice_printer()



